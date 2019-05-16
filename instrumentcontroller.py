import time
import openpyxl
import visa

from os import listdir
from os.path import isfile, join

from PyQt5.QtCore import QObject

# from analyzer import Analyzer
# from analyzermock import AnalyzerMock
# from generator import Generator
# from generatormock import GeneratorMock
# from source import Source
# from sourcemock import SourceMock

# MOCK
mock_enabled = True


class Gen:
    addr = 'gen addr'
    model = 'Gen'
    status = f'{model} at {addr}'
    def __str__(self):
        return 'Gen'


class An:
    addr = 'an addr'
    model = 'An'
    status = f'{model} at {addr}'
    def __str__(self):
        return 'An'


class InstrumentFactory:
    def __init__(self, addr, label):
        self.applicable = None
        self.addr = addr
        self.label = label
        # self.rm = visa.ResourceManager()
    def find(self):
        # TODO remove applicable instrument when found one if needed more than one instrument of the same type
        # TODO: idea: pass list of applicable instruments to differ from the model of the same type?
        instr = self.from_address()
        if not instr:
            return self.try_find()
        return instr
    def from_address(self):
        raise NotImplementedError()
    def try_find(self):
        raise NotImplementedError


class GeneratorFactory(InstrumentFactory):
    def __init__(self, addr):
        super().__init__(addr=addr, label='Генератор')
        self.applicable = ['E4438C', 'N5181B', 'N5183A']
    def from_address(self):
        return Gen()
    def try_find(self):
        return Gen()


class AnalyzerFactory(InstrumentFactory):
    def __init__(self, addr):
        super().__init__(addr=addr, label='Анализатор')
        self.applicable = ['Obzor 304', 'DSA90604A']
    def from_address(self):
        return An()
    def try_find(self):
        return An()


class InstrumentController(QObject):

    def __init__(self, parent=None):
        super().__init__(parent=parent)

        self.requiredInstruments = {
            'Анализатор': AnalyzerFactory('GPIB0::10::INSTR'),
            'Генератор1': GeneratorFactory('GPIB0::5::INSTR'),
            # 'Генератор2': GeneratorFactory('GPIB0::5::INSTR')
        }

        self._instruments = {}

    def __str__(self):
        return f'{self._instruments}'

    def connect(self, addrs):
        print(f'searching for {addrs}')
        for k, v in addrs.items():
            self.requiredInstruments[k].addr = v
        return self._find()

    def _find(self):
        self._instruments = {
            k: v.find() for k, v in self.requiredInstruments.items()
        }
        return all(self._instruments.values())

    @property
    def status(self):
        return [i.status for i in self._instruments.values()]

    def checkSample(self):
        print('instrument manager: check sample')

        if isfile('settings.ini'):
            with open('settings.ini') as f:
                line = f.readline()
                self.pow_limit = float(line.split('=')[1].strip())

        # TODO: report proper condition
        return True

    def checkTaskTable(self):

        def getFileList(data_path):
            return [l for l in listdir(data_path) if isfile(join(data_path, l)) and self.excel_extension in l]

        # TODO: extract measurement manager class
        print('instrument manager: check task table')

        files = getFileList('.')
        length = len(files)
        if length > 1:
            print('too many task tables, abort')
            return False
        elif length <= 0:
            print('no task table found, abort')
            return False

        self._fileName = files[0]
        print('using task table:', self._fileName)

        wb = openpyxl.load_workbook(self._fileName)
        ws = wb.active

        cols = ws.max_column
        rows = ws.max_row
        letters = [chr(ch_code).upper() for ch_code in range(ord('a'), ord('z') + 1)]
        start_col = 2

        count = int((rows - 1) / 3)

        # TODO: validate table
        for row in range(count):
            self._params[row + 1] = [[ws[char + str(row * 3 + 1 + num)].value for num in range(1, 4)] for char in letters[start_col:cols]]

        headers = ['№'] + [ws[char + '1'].value for char in letters[start_col:cols]]

        for model in self.measureModels.values():
            model.initHeader(headers)

        wb.close()
        print('done read')
        return True

    # TODO implement proper measurement algorithm
    def measure(self, letter):
        print(f'start measure letter={letter}')

        self.measureTask(letter)

        self.measureModels[letter].initModel(self._params[letter])

        print('end measure')

    def generator_freq_sweep(self, generator, letter):

        def do_branch(letter, branch, step):

            self.control_voltage = self.step_voltage[step]

            if branch == 1:
                chan1voltage = self.control_voltage
                chan2voltage = 0.0
            else:
                chan1voltage = 0.0
                chan2voltage = self.control_voltage

            self._source.set_voltage(1, chan1voltage, 'V')
            self._source.set_voltage(2, chan2voltage, 'V')
            self._source.set_output(1, 'ON')
            self._source.set_output(2, 'ON')

            generator.set_output('ON')

            pows = list()

            for freq in self.measure_freq[letter].values():
                generator.set_freq(freq, self.measure_freq_unit)
                self._analyzer.set_measure_center_freq(freq, self.measure_freq_unit)
                self._analyzer.set_marker1_x_center(freq, self.measure_freq_unit)

                pows.append(self._analyzer.read_pow(1))
                if not mock_enabled:
                    time.sleep(0.3)

            generator.set_output('OFF')
            self._source.set_output(1, 'OFF')
            self._source.set_output(2, 'OFF')

            return pows

        generator.set_pow(self.measure_pow[1]['p1'], self.measure_pow_unit)

        pows = list()
        for cycle in range(3):
            pows = [do_branch(letter, branch, cycle) for branch in [1, 2]]
        return pows

    def measureTask(self, letter: int):

        print(f'measurement task run, letter={letter}')

        self._generator1.set_modulation('OFF')
        self._generator2.set_modulation('OFF')

        self._analyzer.set_autocalibrate('OFF')
        self._analyzer.set_span(10, 'MHz')

        self._source.set_current(1, self.measure_current, 'mA')
        self._source.set_current(2, self.measure_current, 'mA')

        self._analyzer.set_marker_mode(1, 'POS')

        # gen 1 freq sweep
        pows = self.generator_freq_sweep(self._generator1, letter)
        print(pows)

        # gen 2 freq sweep
        pows = self.generator_freq_sweep(self._generator2, letter)
        print(pows)

        # gen 1 pow sweep
        freq = self.measure_freq[letter]['f3']
        freq_unit = self.measure_freq_unit
        pow_unit = self.measure_pow_unit

        self._generator1.set_freq(freq, freq_unit)

        self._analyzer.set_autocalibrate('OFF')
        self._analyzer.set_span(1, 'MHz')
        self._analyzer.set_marker_mode(1, 'POS')
        self._analyzer.set_measure_center_freq(freq, 'MHz')
        self._analyzer.set_marker1_x_center(freq, 'MHz')

        self._source.set_voltage(1, 3, 'V')
        self._source.set_voltage(2, 0, 'V')
        self._source.set_output(1, 'ON')
        self._source.set_output(2, 'ON')

        self._generator1.set_output('ON')

        for pow in self.measure_pow[letter].values():
            self._generator1.set_pow(pow, pow_unit)

            if not mock_enabled:
                time.sleep(0.3)

            read_pow = self._analyzer.read_pow(1)
            print(read_pow)

        self._generator1.set_output('OFF')

        self._source.set_output(1, 'OFF')
        self._source.set_output(2, 'OFF')
        self._source.set_system_local()

        self._analyzer.set_autocalibrate('ON')